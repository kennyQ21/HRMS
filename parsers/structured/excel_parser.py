import logging
from typing import Any, Dict, List

import openpyxl
import pandas as pd

from ..base import BaseParser

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    def __init__(self):
        super().__init__()

    def parse(self, file_path: str) -> Dict[str, Any]:
        """Parse all sheets of an Excel file and merge into a single dataset."""
        try:
            return self._parse_with_pandas(file_path)
        except Exception as e:
            logger.warning("pandas Excel parse failed (%s), falling back to openpyxl", e)
            return self._parse_with_openpyxl(file_path)

    def validate(self, data: Dict[str, Any]) -> bool:
        """Validate the parsed Excel data."""
        if not data or "data" not in data:
            return False
        required_metadata = ["columns", "rows"]
        return all(key in data["metadata"] for key in required_metadata)

    # ── pandas (preferred) ────────────────────────────────────────────────────

    def _parse_with_pandas(self, file_path: str) -> Dict[str, Any]:
        """
        Read every sheet.
        Returns a combined record list with a '__sheet__' column so callers
        know which sheet each row came from.
        """
        all_sheets: Dict[str, Any] = pd.read_excel(file_path, sheet_name=None)  # {name: df}

        combined_rows: List[Dict] = []
        all_columns: list = []
        sheet_summary: List[Dict] = []

        for sheet_name, df in all_sheets.items():
            df["__sheet__"] = sheet_name
            records = df.to_dict("records")
            combined_rows.extend(records)

            # Track per-sheet column names (excluding the injected __sheet__)
            sheet_cols = [c for c in df.columns.tolist() if c != "__sheet__"]
            for col in sheet_cols:
                if col not in all_columns:
                    all_columns.append(col)

            sheet_summary.append({"sheet": sheet_name, "rows": len(df), "columns": sheet_cols})
            logger.debug("Parsed sheet '%s': %d rows, %d columns", sheet_name, len(df), len(sheet_cols))

        return {
            "data": combined_rows,
            "metadata": {
                "columns": all_columns,
                "rows": len(combined_rows),
                "sheets": sheet_summary,
                "parser": "pandas",
            },
        }

    # ── openpyxl fallback ─────────────────────────────────────────────────────

    def _parse_with_openpyxl(self, file_path: str) -> Dict[str, Any]:
        """Read every sheet using openpyxl."""
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        combined_rows: List[Dict] = []
        all_columns: list = []
        sheet_summary: List[Dict] = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows_iter = sheet.iter_rows(values_only=True)

            try:
                headers = [cell if cell is not None else f"col_{i}" for i, cell in enumerate(next(rows_iter))]
            except StopIteration:
                # Empty sheet
                sheet_summary.append({"sheet": sheet_name, "rows": 0, "columns": []})
                continue

            sheet_rows: List[Dict] = []
            for row in rows_iter:
                row_dict = {h: v for h, v in zip(headers, row)}
                row_dict["__sheet__"] = sheet_name
                sheet_rows.append(row_dict)

            combined_rows.extend(sheet_rows)

            for col in headers:
                if col not in all_columns:
                    all_columns.append(col)

            sheet_summary.append({"sheet": sheet_name, "rows": len(sheet_rows), "columns": headers})
            logger.debug("Parsed sheet '%s': %d rows", sheet_name, len(sheet_rows))

        wb.close()

        return {
            "data": combined_rows,
            "metadata": {
                "columns": all_columns,
                "rows": len(combined_rows),
                "sheets": sheet_summary,
                "parser": "openpyxl",
            },
        }